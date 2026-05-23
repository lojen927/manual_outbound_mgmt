from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, DetailView, UpdateView, View, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.utils.translation import gettext
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django import forms
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote

from .models import ManualOutboundOrder, ManualOutboundItem, Reason, Material
from .forms import (
    ManualOutboundOrderForm,
    ManualOutboundItemFormSet,
    ManualOutboundItemCloseForm,
    ManualOutboundVoidForm,
    ReasonForm,
)


class ManualOutboundListView(LoginRequiredMixin, ListView):
    model = ManualOutboundOrder
    template_name = 'outbound/list.html'
    paginate_by = 15

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get('q', '').strip()
        material_code = self.request.GET.get('material_code', '').strip()
        status_filter = self.request.GET.get('status', '').strip()
        needs_posting = self.request.GET.get('needs_system_posting', '').strip()
        closed_status = self.request.GET.get('closed_status', '').strip()
        date_from = self.request.GET.get('date_from', '').strip()
        date_to = self.request.GET.get('date_to', '').strip()

        if q:
            qs = qs.filter(
                Q(order_no__icontains=q) |
                Q(department__icontains=q) |
                Q(applicant__icontains=q) |
                Q(employee_id__icontains=q)
            )
        if material_code:
            qs = qs.filter(items__material_code__icontains=material_code).distinct()
        if status_filter in ('draft', 'submitted', 'completed', 'voided'):
            qs = qs.filter(status=status_filter)
        if needs_posting in ('true', 'false'):
            qs = qs.filter(needs_system_posting=(needs_posting == 'true'))
        if closed_status == 'all_closed':
            qs = qs.annotate(open_count=Count('items', filter=Q(items__status='open'))).filter(open_count=0)
        elif closed_status == 'has_open':
            qs = qs.annotate(open_count=Count('items', filter=Q(items__status='open'))).filter(open_count__gt=0)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        return qs.select_related('created_by')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['params'] = self.request.GET.copy()
        for order in ctx['object_list']:
            order.open_count = order.items.filter(status='open').count()
            order.total_count = order.items.count()
        return ctx


class ManualOutboundCreateView(LoginRequiredMixin, CreateView):
    model = ManualOutboundOrder
    form_class = ManualOutboundOrderForm
    template_name = 'outbound/form.html'
    success_url = reverse_lazy('outbound_list')

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = _('新增手工出库单')
        if self.request.POST:
            ctx['item_formset'] = ManualOutboundItemFormSet(self.request.POST)
        else:
            ctx['item_formset'] = ManualOutboundItemFormSet()
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        item_formset = ctx['item_formset']
        if not item_formset.is_valid():
            return self.render_to_response(ctx)
        self.object = form.save(commit=False)
        self.object.created_by = self.request.user
        self.object.status = 'draft'
        self.object.save()
        item_formset.instance = self.object
        item_formset.save()
        messages.success(self.request, gettext('手工出库单 %(order_no)s 草稿保存成功。') % {'order_no': self.object.order_no})
        return redirect(self.success_url)

    def form_invalid(self, form):
        ctx = self.get_context_data(form=form)
        return self.render_to_response(ctx)


class ManualOutboundUpdateView(LoginRequiredMixin, UpdateView):
    model = ManualOutboundOrder
    form_class = ManualOutboundOrderForm
    template_name = 'outbound/form.html'

    def dispatch(self, request, *args, **kwargs):
        order = self.get_object()
        if order.status == 'submitted' and not request.user.is_superuser:
            messages.error(request, gettext('已提交的出库单仅限管理员编辑。'))
            return redirect('outbound_detail', pk=order.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('outbound_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = _('编辑出库单 - %(order_no)s') % {'order_no': self.object.order_no}
        if self.request.POST:
            ctx['item_formset'] = ManualOutboundItemFormSet(
                self.request.POST, instance=self.object
            )
        else:
            ctx['item_formset'] = ManualOutboundItemFormSet(instance=self.object)
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        item_formset = ctx['item_formset']
        if not item_formset.is_valid():
            return self.render_to_response(ctx)
        self.object = form.save()
        item_formset.instance = self.object
        item_formset.save()
        messages.success(self.request, gettext('出库单 %(order_no)s 已更新。') % {'order_no': self.object.order_no})
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        ctx = self.get_context_data(form=form)
        return self.render_to_response(ctx)


class ManualOutboundSubmitView(LoginRequiredMixin, View):
    def post(self, request, pk):
        order = get_object_or_404(ManualOutboundOrder, pk=pk)
        if order.status != 'draft':
            messages.warning(request, gettext('只有草稿状态的出库单可以提交。'))
            return redirect('outbound_detail', pk=pk)
        order.status = 'submitted'
        order.save()
        messages.success(request, gettext('出库单 %(order_no)s 已提交。') % {'order_no': order.order_no})
        return redirect('outbound_detail', pk=pk)


class ManualOutboundDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        order = get_object_or_404(ManualOutboundOrder, pk=pk)
        if order.status != 'draft' and not request.user.is_superuser:
            messages.error(request, gettext('仅管理员有权限删除非草稿状态的出库单。'))
            return redirect('outbound_detail', pk=pk)
        order_no = order.order_no
        order.delete()
        messages.success(request, gettext('出库单 %(order_no)s 已删除。') % {'order_no': order_no})
        return redirect('outbound_list')


class ManualOutboundDetailView(LoginRequiredMixin, DetailView):
    model = ManualOutboundOrder
    template_name = 'outbound/detail.html'

    def get_queryset(self):
        return super().get_queryset().select_related('created_by').prefetch_related('items')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['close_form'] = ManualOutboundItemCloseForm()
        ctx['open_items'] = self.object.items.filter(status='open')
        return ctx


class ManualOutboundItemCloseView(LoginRequiredMixin, View):
    def post(self, request, order_pk, item_pk):
        order = get_object_or_404(ManualOutboundOrder, pk=order_pk)
        item = get_object_or_404(ManualOutboundItem, pk=item_pk, order=order)

        if item.status == 'closed':
            messages.warning(request, gettext('该项已关闭，无需重复操作。'))
            return redirect('outbound_detail', pk=order_pk)

        form = ManualOutboundItemCloseForm(request.POST)
        if not form.is_valid():
            for field, errors in form.errors.items():
                for err in errors:
                    messages.error(request, f'{err}')
            return redirect('outbound_detail', pk=order_pk)

        mode = form.cleaned_data['close_mode']
        item.status = 'closed'
        item.closed_at = timezone.now()
        item.closed_by = request.user

        if mode == 'system':
            item.system_order_no = form.cleaned_data['system_order_no']
        else:
            item.close_reason = form.cleaned_data['close_reason']

        item.save()
        # Auto-complete order when all items are closed
        if order.items.filter(status='open').count() == 0:
            order.status = 'completed'
            order.save()
        messages.success(request, gettext('明细项 %(code)s 已关闭。') % {'code': item.material_code})
        return redirect('outbound_detail', pk=order_pk)


class ManualOutboundPrintView(LoginRequiredMixin, DetailView):
    model = ManualOutboundOrder
    template_name = 'outbound/print.html'

    def get_queryset(self):
        return super().get_queryset().prefetch_related('items')

    def dispatch(self, request, *args, **kwargs):
        order = self.get_object()
        if order.status == 'voided':
            messages.error(request, gettext('已作废的出库单不能打印。'))
            return redirect('outbound_detail', pk=order.pk)
        if order.status == 'completed' and not request.user.is_superuser:
            messages.error(request, gettext('已完成的出库单仅限管理员打印。'))
            return redirect('outbound_detail', pk=order.pk)
        return super().dispatch(request, *args, **kwargs)


class ManualOutboundExportView(LoginRequiredMixin, View):
    def get(self, request):
        qs = ManualOutboundOrder.objects.exclude(status='draft').select_related('created_by')
        q = request.GET.get('q', '').strip()
        material_code = request.GET.get('material_code', '').strip()
        status_filter = request.GET.get('status', '').strip()
        needs_posting = request.GET.get('needs_system_posting', '').strip()
        closed_status = request.GET.get('closed_status', '').strip()
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()

        if q:
            qs = qs.filter(
                Q(order_no__icontains=q) |
                Q(department__icontains=q) |
                Q(applicant__icontains=q) |
                Q(employee_id__icontains=q)
            )
        if material_code:
            qs = qs.filter(items__material_code__icontains=material_code).distinct()
        if status_filter in ('draft', 'submitted', 'completed', 'voided'):
            qs = qs.filter(status=status_filter)
        if needs_posting in ('true', 'false'):
            qs = qs.filter(needs_system_posting=(needs_posting == 'true'))
        if closed_status == 'all_closed':
            qs = qs.annotate(open_count=Count('items', filter=Q(items__status='open'))).filter(open_count=0)
        elif closed_status == 'has_open':
            qs = qs.annotate(open_count=Count('items', filter=Q(items__status='open'))).filter(open_count__gt=0)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = gettext('手工出库记录')

        headers = [
            gettext('出库单号'), gettext('单据状态'), gettext('创建时间'),
            gettext('领料部门'), gettext('领料人'), gettext('工号'), gettext('联系电话'),
            gettext('手工领料原因'), gettext('备注说明'), gettext('需补过账'), gettext('未审批单号'),
            gettext('物料编码'), gettext('物料描述'), gettext('规格'), gettext('数量'), gettext('单位'), gettext('存货库位'),
            gettext('明细备注'), gettext('明细状态'), gettext('补录系统出库单号'), gettext('关闭原因'),
            gettext('关闭时间'), gettext('关闭人'),
        ]

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        row = 2
        for order in qs:
            items = order.items.all()
            if not items:
                data = self._order_to_row(order, None)
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                row += 1
            else:
                for item in items:
                    data = self._order_to_row(order, item)
                    for col, val in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.border = thin_border
                    row += 1

        col_widths = [18, 10, 16, 12, 10, 10, 14, 16, 24, 10, 20, 14, 20, 12, 10, 8, 12, 16, 8, 18, 20, 18, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        filename = gettext('出库单列表_%(date)s.xlsx') % {'date': timezone.localdate().strftime("%y%m%d")}
        ascii_name = f'outbound_{timezone.localdate().strftime("%y%m%d")}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"
        wb.save(response)
        return response

    @staticmethod
    def _order_to_row(order, item):
        row_data = [
            order.order_no, order.get_status_display(),
            order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '',
            order.department, order.applicant,
            order.employee_id, order.phone, order.reason_label,
            order.remark,
            gettext('是') if order.needs_system_posting else gettext('否'),
            order.pending_order_numbers,
        ]
        if item:
            row_data.extend([
                item.material_code, item.material_name, item.specification,
                float(item.quantity), item.unit, item.storage_location,
                item.remark, item.get_status_display(),
                item.system_order_no, item.close_reason,
                item.closed_at.strftime('%Y-%m-%d %H:%M') if item.closed_at else '',
                item.closed_by.username if item.closed_by else '',
            ])
        else:
            row_data.extend([''] * 12)
        return row_data


class ManualOutboundVoidView(LoginRequiredMixin, FormView):
    """作废已提交的出库单（提交后2小时内有效）。"""
    form_class = ManualOutboundVoidForm
    template_name = 'outbound/void_confirm.html'

    def dispatch(self, request, *args, **kwargs):
        order = get_object_or_404(ManualOutboundOrder, pk=self.kwargs['pk'])
        if order.status != 'submitted':
            messages.error(request, gettext('只有已提交状态的出库单可以作废。'))
            return redirect('outbound_detail', pk=order.pk)

        elapsed = timezone.now() - order.updated_at
        if elapsed.total_seconds() > 7200:  # 2 hours
            messages.error(request, gettext('出库单 %(order_no)s 已提交超过2小时，无法作废。') % {'order_no': order.order_no})
            return redirect('outbound_detail', pk=order.pk)

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['order'] = get_object_or_404(ManualOutboundOrder, pk=self.kwargs['pk'])
        ctx['title'] = _('作废出库单')
        return ctx

    def form_valid(self, form):
        order = get_object_or_404(ManualOutboundOrder, pk=self.kwargs['pk'])
        order.status = 'voided'
        order.needs_system_posting = False
        order.voided_by = self.request.user
        order.voided_at = timezone.now()
        order.void_reason = form.cleaned_data['void_reason']
        order.save()

        # Close all open items with reason "出库单已作废"
        now = timezone.now()
        for item in order.items.filter(status='open'):
            item.status = 'closed'
            item.closed_at = now
            item.closed_by = self.request.user
            item.close_reason = gettext('出库单已作废')
            item.save()

        messages.success(self.request, gettext('出库单 %(order_no)s 已作废。') % {'order_no': order.order_no})
        return redirect('outbound_detail', pk=order.pk)


class MaterialLookupView(LoginRequiredMixin, View):
    def get(self, request):
        q = request.GET.get('q', '').strip()
        if not q:
            return JsonResponse({'found': False})
        try:
            material = Material.objects.get(material_code=q)
            return JsonResponse({
                'found': True,
                'material_name': material.material_name,
                'specification': material.specification,
                'unit': material.unit,
            })
        except Material.DoesNotExist:
            return JsonResponse({'found': False})


# ---- Admin-only mixin ----

class SuperuserRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser

    def handle_no_permission(self):
        messages.error(self.request, gettext('仅管理员可访问此页面。'))
        return redirect('outbound_list')


# ---- User Management Views ----

class UserListView(SuperuserRequiredMixin, LoginRequiredMixin, View):
    def get(self, request):
        users = User.objects.all().order_by('-is_superuser', 'username')
        return render(request, 'outbound/user_list.html', {'users': users})


class UserCreateForm(forms.ModelForm):
    password = forms.CharField(
        label=_('密码'), max_length=128,
        widget=forms.PasswordInput(attrs={'class': 'form-control'}),
    )
    password_confirm = forms.CharField(
        label=_('确认密码'), max_length=128,
        widget=forms.PasswordInput(attrs={'class': 'form-control'}),
    )

    class Meta:
        model = User
        fields = ['username', 'email', 'first_name', 'is_superuser', 'is_active']
        labels = {
            'username': _('用户名'),
            'email': _('邮箱'),
            'first_name': _('姓名'),
            'is_superuser': _('超级用户'),
            'is_active': _('活跃'),
        }
        widgets = {
            'username': forms.TextInput(attrs={'class': 'form-control'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'first_name': forms.TextInput(attrs={'class': 'form-control'}),
            'is_superuser': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }

    def clean(self):
        cleaned = super().clean()
        pwd = cleaned.get('password')
        confirm = cleaned.get('password_confirm')
        if pwd and confirm and pwd != confirm:
            self.add_error('password_confirm', gettext('两次输入的密码不一致。'))
        return cleaned

    def save(self, commit=True):
        user = super().save(commit=False)
        user.set_password(self.cleaned_data['password'])
        if commit:
            user.save()
        return user


class UserCreateView(SuperuserRequiredMixin, LoginRequiredMixin, FormView):
    template_name = 'outbound/user_form.html'
    form_class = UserCreateForm
    success_url = reverse_lazy('user_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = _('新增用户')
        return ctx

    def form_valid(self, form):
        form.save()
        messages.success(self.request, gettext('用户 %(username)s 创建成功。') % {'username': form.cleaned_data["username"]})
        return super().form_valid(form)


class UserPasswordResetForm(forms.Form):
    password = forms.CharField(
        label=_('新密码'), max_length=128,
        widget=forms.PasswordInput(attrs={'class': 'form-control'}),
    )
    password_confirm = forms.CharField(
        label=_('确认密码'), max_length=128,
        widget=forms.PasswordInput(attrs={'class': 'form-control'}),
    )

    def clean(self):
        cleaned = super().clean()
        pwd = cleaned.get('password')
        confirm = cleaned.get('password_confirm')
        if pwd and confirm and pwd != confirm:
            self.add_error('password_confirm', gettext('两次输入的密码不一致。'))
        return cleaned


class UserPasswordResetView(SuperuserRequiredMixin, LoginRequiredMixin, FormView):
    template_name = 'outbound/user_form.html'
    form_class = UserPasswordResetForm
    success_url = reverse_lazy('user_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = get_object_or_404(User, pk=self.kwargs['pk'])
        ctx['title'] = gettext('重置密码 - %(username)s') % {'username': user.username}
        ctx['target_user'] = user
        return ctx

    def form_valid(self, form):
        user = get_object_or_404(User, pk=self.kwargs['pk'])
        user.set_password(form.cleaned_data['password'])
        user.save()
        messages.success(self.request, gettext('用户 %(username)s 密码已重置。') % {'username': user.username})
        return super().form_valid(form)


class UserToggleActiveView(SuperuserRequiredMixin, LoginRequiredMixin, View):
    def get(self, request, pk):
        if pk == request.user.pk:
            messages.warning(request, gettext('不能禁用自己。'))
            return redirect('user_list')
        user = get_object_or_404(User, pk=pk)
        user.is_active = not user.is_active
        user.save()
        status = gettext('启用') if user.is_active else gettext('禁用')
        messages.success(request, gettext('用户 %(username)s 已%(status)s。') % {'username': user.username, 'status': status})
        return redirect('user_list')


# ---- Reason Management Views ----

class ReasonListView(SuperuserRequiredMixin, LoginRequiredMixin, ListView):
    model = Reason
    template_name = 'outbound/reason_list.html'
    paginate_by = 20


class ReasonCreateView(SuperuserRequiredMixin, LoginRequiredMixin, CreateView):
    model = Reason
    form_class = ReasonForm
    template_name = 'outbound/reason_form.html'
    success_url = reverse_lazy('reason_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = _('新增领料原因')
        return ctx

    def form_valid(self, form):
        messages.success(self.request, gettext('领料原因 "%(label)s" 创建成功。') % {'label': form.instance.label})
        return super().form_valid(form)


class ReasonUpdateView(SuperuserRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Reason
    form_class = ReasonForm
    template_name = 'outbound/reason_form.html'
    success_url = reverse_lazy('reason_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = _('编辑领料原因 - %(label)s') % {'label': self.object.label}
        return ctx

    def form_valid(self, form):
        messages.success(self.request, gettext('领料原因 "%(label)s" 已更新。') % {'label': form.instance.label})
        return super().form_valid(form)


class ReasonDeleteView(SuperuserRequiredMixin, LoginRequiredMixin, View):
    def post(self, request, pk):
        reason = get_object_or_404(Reason, pk=pk)
        label = reason.label
        # Check if any order references this reason code
        if ManualOutboundOrder.objects.filter(reason=reason.code).exists():
            messages.error(
                request, gettext('领料原因 "%(label)s" 已被出库单使用，无法删除。请先将其设为禁用。') % {'label': label}
            )
        else:
            reason.delete()
            messages.success(request, gettext('领料原因 "%(label)s" 已删除。') % {'label': label})
        return redirect('reason_list')
