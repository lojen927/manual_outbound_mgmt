import os
from datetime import timedelta
from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils import timezone
from django.db.models import Count, Q
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from outbound.models import ManualOutboundOrder


class Command(BaseCommand):
    help = '导出非草稿状态的所有出库单到 Excel 文件'

    def add_arguments(self, parser):
        parser.add_argument(
            '--days', type=int, default=0,
            help='仅导出最近 N 天的记录（默认全部）',
        )
        parser.add_argument(
            '--output', type=str, default='',
            help='输出目录（默认 Downloads）',
        )

    def handle(self, *args, **options):
        qs = ManualOutboundOrder.objects.exclude(status='draft').select_related('created_by')

        days = options['days']
        if days > 0:
            since = timezone.now() - timedelta(days=days)
            qs = qs.filter(created_at__gte=since)

        output_dir = options['output']
        if not output_dir:
            output_dir = str(Path.home() / 'Downloads')

        os.makedirs(output_dir, exist_ok=True)

        today = timezone.localdate()
        filename = f'出库单列表_{today.strftime("%y%m%d")}.xlsx'
        filepath = os.path.join(output_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '手工出库记录'

        headers = [
            '出库单号', '单据状态', '创建时间',
            '领料部门', '领料人', '工号', '联系电话',
            '手工领料原因', '备注说明', '需补过账', '未审批单号',
            '物料编码', '物料描述', '规格', '数量', '单位', '存货库位',
            '明细备注', '明细状态', '补录系统出库单号', '关闭原因',
            '关闭时间', '关闭人',
        ]

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        row = 2
        for order in qs:
            items = order.items.all()
            if not items:
                data = self._order_to_row(order, None)
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                row += 1
            else:
                for item in items:
                    data = self._order_to_row(order, item)
                    for col, val in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.border = thin_border
                    row += 1

        col_widths = [18, 10, 16, 12, 10, 10, 14, 16, 24, 10, 20, 14, 20, 12, 10, 8, 12, 16, 8, 18, 20, 18, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(filepath)
        self.stdout.write(self.style.SUCCESS(f'已导出: {filepath}'))

    @staticmethod
    def _order_to_row(order, item):
        row_data = [
            order.order_no, order.get_status_display(),
            order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '',
            order.department, order.applicant,
            order.employee_id, order.phone, order.reason_label,
            order.remark,
            '是' if order.needs_system_posting else '否',
            order.pending_order_numbers,
        ]
        if item:
            row_data.extend([
                item.material_code, item.material_name, item.specification,
                float(item.quantity), item.unit, item.storage_location,
                item.remark, item.get_status_display(),
                item.system_order_no, item.close_reason,
                item.closed_at.strftime('%Y-%m-%d %H:%M') if item.closed_at else '',
                item.closed_by.username if item.closed_by else '',
            ])
        else:
            row_data.extend([''] * 12)
        return row_data
